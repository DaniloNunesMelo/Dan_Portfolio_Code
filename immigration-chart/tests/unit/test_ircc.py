"""Unit tests for src/fetchers/ircc.py"""
from io import BytesIO

import pytest

from src.fetchers.ircc import IRCCFetcher
from src.fetchers.base import FetchError

CANONICAL_COLUMNS = [
    "ref_area", "ref_area_name", "counterpart", "counterpart_name",
    "time_period", "year", "quarter", "var_code", "metric",
    "sex", "gender", "area_name", "reg_name", "province",
    "imm_category", "obs_value", "obs_status", "source_dataset", "fetch_ts",
]


# ── _parse_ircc_xlsx ──────────────────────────────────────────────────────────

def test_parse_ircc_xlsx_canonical_schema(ircc_xlsx_bytes):
    fetcher = IRCCFetcher()
    df = fetcher._parse_ircc_xlsx(ircc_xlsx_bytes, "citizenship")
    for col in CANONICAL_COLUMNS:
        assert col in df.columns, f"Missing column: {col}"


def test_parse_ircc_xlsx_var_code_pr(ircc_xlsx_bytes):
    fetcher = IRCCFetcher()
    df = fetcher._parse_ircc_xlsx(ircc_xlsx_bytes, "citizenship")
    assert (df["var_code"] == "PR").all()


def test_parse_ircc_xlsx_source_tag(ircc_xlsx_bytes):
    fetcher = IRCCFetcher()
    df = fetcher._parse_ircc_xlsx(ircc_xlsx_bytes, "citizenship")
    assert (df["source_dataset"] == "IRCC").all()


def test_parse_ircc_xlsx_ref_area(ircc_xlsx_bytes):
    fetcher = IRCCFetcher()
    df = fetcher._parse_ircc_xlsx(ircc_xlsx_bytes, "citizenship")
    assert (df["ref_area"] == "CAN").all()


def test_parse_ircc_xlsx_filters_zero_values(ircc_xlsx_bytes):
    fetcher = IRCCFetcher()
    df = fetcher._parse_ircc_xlsx(ircc_xlsx_bytes, "citizenship")
    assert (df["obs_value"] > 0).all()


def test_parse_ircc_xlsx_excludes_footnote_rows(ircc_xlsx_bytes):
    fetcher = IRCCFetcher()
    df = fetcher._parse_ircc_xlsx(ircc_xlsx_bytes, "citizenship")
    # "Source: IRCC" row should not appear as counterpart_name
    assert not any("Source" in str(v) for v in df["counterpart_name"].tolist())


def test_parse_ircc_xlsx_no_annual_cols_raises():
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    for i in range(6):
        ws.append(["row"] * 5)
    # Row 3 (index 3) has no "YYYY Total" pattern
    ws.cell(row=4, column=2, value="Q1 2020")  # wrong pattern

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    fetcher = IRCCFetcher()
    with pytest.raises(FetchError, match="No annual total columns"):
        fetcher._parse_ircc_xlsx(buf, "citizenship")


# ── _to_canonical province mode ───────────────────────────────────────────────

def test_to_canonical_province_mode(ircc_xlsx_bytes):
    """province_cat file_key should populate the province column."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Title"])
    ws.append([""])
    ws.append(["Category", 2020, None])
    ws.append(["Category", "2020 Total", "Q1"])
    ws.append(["Category", "Jan", "Feb"])
    ws.append(["Ontario", 10000, 2500])
    ws.append(["Quebec", 5000, 1200])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    fetcher = IRCCFetcher()
    df = fetcher._parse_ircc_xlsx(buf, "province_cat")
    # province should be mapped to ISO codes
    assert df["province"].notna().any()
    assert "CA-ON" in df["province"].tolist() or "CA-QC" in df["province"].tolist()


def test_to_canonical_citizenship_mode(ircc_xlsx_bytes):
    """citizenship file_key should set counterpart from country names."""
    fetcher = IRCCFetcher()
    df = fetcher._parse_ircc_xlsx(ircc_xlsx_bytes, "citizenship")
    # province should be None for citizenship mode
    assert df["province"].isna().all() or (df["province"] == None).all()  # noqa: E711
